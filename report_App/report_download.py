# This is a sample Python script.

# Press Shift+F10 to execute it or replace it with your code.
# Press Double Shift to search everywhere for classes, files, tool windows, actions, and settings.
import requests
import datetime
import xlsxwriter
from openpyxl import Workbook, load_workbook
from openpyxl.chart import (
    Reference,
    Series,
    BarChart3D,
)

class BearerAuth(requests.auth.AuthBase):
    def __init__(self, token):
        self.token = token
    def __call__(self, r):
        r.headers["authorization"] = "Bearer " + self.token
        return r

def web_chat(url_val,bear_auth,month_check,room_data_get):
    response = requests.get(url_val, auth=BearerAuth(bear_auth))
    print(response)
    data = response.json()
    print(data['items'])

    get_date_list = []
    get_Id_list = []
    get_parentID_list = []

    for i in range(len(data['items'])):
        k=str(data['items'][i]['created'])
        month_val = k.split('-')
        year = month_val[0]
        if month_val[1] == str(month_check):
            get_date_list.append(data['items'][i]['created'])
            get_Id_list.append(data['items'][i]['id'])
            if "parentId" in data['items'][i]:
                get_parentID_list.append(data['items'][i]['parentId'])

    print(get_Id_list)
    get_date_list.reverse()
    get_parentID_list.reverse()
    print(get_Id_list)

    get_Id_list.reverse()
    print(get_Id_list)

    month_number = str(month_check)
    datetime_object = datetime.datetime.strptime(month_number, "%m")
    month_name = datetime_object.strftime("%b")

    workbook = xlsxwriter.Workbook('folder_name/'+room_data_get+"_"+month_name+str(year)+'.xlsx')
    workbook.add_worksheet("Result")
    worksheet = workbook.add_worksheet(month_name+str(year))

    Dict_ans_email = {}

    row = 0
    col = 0
    heading_list = ['Date', 'Email of question', 'Domain', 'Question asked', 'Room', 'Answered by', 'Answer Given']
    count = 0
    for title in (heading_list):
        worksheet.write(row, col + count, title)
        count = count + 1
    row += 1

    for k in get_Id_list:
        for j in range(len(data['items'])):
            if  k == data['items'][j]['id']:
                print("=============================================================")
                print("------------------Question------------------------")
                print(data['items'][j]['created'])
                split_date = str(data['items'][j]['created'])
                date_val = split_date.split('T')
                print(date_val[0])
                worksheet.write(row,col, date_val[0])
                print(data['items'][j]['personEmail'])
                worksheet.write(row,col+1, data['items'][j]['personEmail'])
                worksheet.write(row,col+2, " ")
                if "text" in data['items'][j]:
                    print(data['items'][j]['text'])
                    worksheet.write(row,col+3,data['items'][j]['text'])
                worksheet.write(row,col+4,room_data_get)
                for l in range(len(data['items'])):
                    if "parentId" in data['items'][l]:
                        if data['items'][l]['parentId'] == k:
                            print("------------------Answer------------------------")
                            if data['items'][l]['personEmail'] in Dict_ans_email.keys():
                                values = Dict_ans_email.get(data['items'][l]['personEmail'])
                                values = values + 1
                                Dict_ans_email[data['items'][l]['personEmail']] = values

                            else:
                                Dict_ans_email[data['items'][l]['personEmail']] = 1
                                print(Dict_ans_email)

                            print(data['items'][l]['personEmail'])
                            worksheet.write(row,col+5,data['items'][l]['personEmail'])
                            if "text" in data['items'][l]:
                                print(data['items'][l]['text'])
                                worksheet.write(row,col+6, data['items'][l]['text'])
                            row += 1


    print(Dict_ans_email)

    reply_count_list = sorted(Dict_ans_email.items(), key=lambda x: x[1])
    reply_count_list.reverse()
    result_list_val = dict(reply_count_list)
    print(result_list_val)

    key_val = result_list_val.keys()
    values_val = result_list_val.values()
    key_value_list = zip(key_val, values_val)

    path_val = 'folder_name/'+room_data_get+"_"+month_name+str(year)+'.xlsx'
    workbook.close()

    chat_met = month_name+str(year)

    barchat_val(path_val,key_value_list,chat_met)
    filename_val = room_data_get+"_"+month_name+str(year)

    return filename_val
    #pymsgbox.alert(filename_val, 'Download!')

# Press the green button in the gutter to run the script.


def download(room_get_data,get_date,get_bear_token):
    room_data = room_get_data
    date = get_date
    datee = datetime.datetime.strptime(date, "%Y-%m-%d")
    month = int(datee.month)-1
    if month<10:
        month="0"+str(month)
    if month=="00":
        month="12"
    print(month)
    url_get_data = ''
    date_format_value = date+'T00:00:00.000Z'
    bear_token = get_bear_token
    if room_data == "support":
        url_get_data ='https://webexapis.com/v1/messages?roomId=Y2lzY29zcGFyazovL3VzL1JPT00vNGZlOTQ2MTAtZjA2MS0xMWU1LWI4Y2UtMTEzZjhkZmMxNGJl&before='+date_format_value+'&max=1000'
    elif room_data == "program":
        url_get_data ='https://webexapis.com/v1/messages?roomId=Y2lzY29zcGFyazovL3VzL1JPT00vYzQ2OTk3NTAtZGIyNy0xMWU1LWI0ZjQtZmJmMjI3Y2ZmYWYz&before='+date_format_value+'&max=1000'

    file_list_name = web_chat(url_get_data,bear_token,month,room_data)

    return file_list_name

def barchat_val(path,get_list_of_key_pair,get_chart_met):

    wb = load_workbook(path)
    ws = wb['chart']

    rows = get_list_of_key_pair
    for row in rows:
        ws.append(row)

    data = Reference(ws, min_col=2, min_row=1, max_col=3, max_row=10)
    titles = Reference(ws, min_col=1, min_row=1, max_row=10)
    chart = BarChart3D()
    chart.title = "Chat metrics"+get_chart_met
    chart.add_data(data=data, titles_from_data=True)
    chart.set_categories(titles)

    ws.add_chart(chart, "E5")
    wb.save(path)

# See PyCharm help at https://www.jetbrains.com/help/pycharm/
